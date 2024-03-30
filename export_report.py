from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from collections import defaultdict

# Example usage
input_file = 'testing.xlsx'
output_file = 'testing_1.xlsx'
sheet_name = '1.1 Chi tiết CCTB'
column_index = 7  # Index of the column to check (e.g., column C)

# Load the workbook
wb = load_workbook(input_file)
# Select the worksheet
ws = wb[sheet_name]

## logic
rows_to_remove = []
dict_npp = defaultdict(list)
for row in ws.iter_rows(min_row=2, max_col=column_index, max_row=ws.max_row):
    print(row[column_index - 1].value)
    print(row[0].row)
    dict_npp[row[column_index - 1].value].append(row[0].row)

for npp_key, value in dict_npp.items():
    npp_wb = wb
    npp_ws = npp_wb[sheet_name]
    print(npp_key)
    for currentKey, listOfRows in dict_npp.items():
        print(currentKey)
        if npp_key != currentKey:
            for rowNumber in listOfRows:
                npp_ws.delete_rows(rowNumber)
        else:
            continue
    # Save the workbook
    npp_wb.save(npp + ".xlsx")
    print(npp + "DONE")


# remove_rows_by_value(input_file, output_file, sheet_name, column_index, value_to_check)
